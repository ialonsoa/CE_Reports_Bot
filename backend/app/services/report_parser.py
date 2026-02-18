"""
Parses uploaded report files (PDF, Word, Excel) and extracts text content
so the AI can learn the format and structure of past reports.
"""
import pdfplumber
import docx
import openpyxl
from pathlib import Path


def parse_pdf(file_path: str) -> str:
    text_blocks = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                text_blocks.append(text.strip())
    return "\n\n".join(text_blocks)


def parse_word(file_path: str) -> str:
    doc = docx.Document(file_path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def parse_excel(file_path: str) -> str:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sections = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cleaned = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cleaned):
                rows.append("\t".join(cleaned))
        if rows:
            sections.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
    return "\n\n".join(sections)


def parse_file(file_path: str) -> tuple[str, str]:
    """
    Returns (content, file_type) for a given file path.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".pdf":
        return parse_pdf(file_path), "pdf"
    elif suffix in (".docx", ".doc"):
        return parse_word(file_path), "word"
    elif suffix in (".xlsx", ".xls"):
        return parse_excel(file_path), "excel"
    elif suffix in (".txt", ".md"):
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read(), "text"
    else:
        raise ValueError(f"Unsupported file type: {suffix}")
